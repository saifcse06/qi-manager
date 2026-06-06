from django.shortcuts import render, redirect
from django.contrib import messages
from django.views import View
from django.views.generic import TemplateView, ListView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Count, F, Case, When, DecimalField
from django.utils import timezone
from datetime import datetime, timedelta
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from clients.models import Client
from quotations.models import Quotation
from invoices.models import Invoice, InvoiceHistory
from payments.models import Payment
from .models import ReportTemplate, SavedReport


def _sidebar_context():
    from accounts.models import User
    return {
        'report_count': SavedReport.objects.count(),
    }


class RoleRequiredMixin(UserPassesTestMixin):
    required_roles = []

    def test_func(self):
        if not self.request.user.is_authenticated:
            return False
        if self.request.user.is_superuser:
            return True
        if not self.required_roles:
            return True
        return self.request.user.roles.filter(name__in=self.required_roles).exists()

    def handle_no_permission(self):
        messages.error(self.request, "You don't have permission to access this page.")
        return redirect('home')


class PermissionRequiredMixin(UserPassesTestMixin):
    required_permission = None

    def test_func(self):
        if not self.request.user.is_authenticated:
            return False
        if self.request.user.is_superuser:
            return True
        if not self.required_permission:
            return True
        return self.request.user.roles.filter(permissions__codename=self.required_permission).exists()

    def handle_no_permission(self):
        messages.error(self.request, "You don't have permission to access this page.")
        return redirect('home')


class ListViewMixin:
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(_sidebar_context())
        return context


class ReportDashboardView(LoginRequiredMixin, RoleRequiredMixin, PermissionRequiredMixin, ListViewMixin, TemplateView):
    template_name = 'reports/dashboard.html'
    required_roles = ['Super Admin', 'Admin']
    required_permission = 'reports.view_reports'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        current_month_start = today.replace(day=1)
        current_year_start = today.replace(month=1, day=1)

        context['today_sales'] = Payment.objects.filter(status='completed', payment_date=today).aggregate(total=Sum('amount'))['total'] or 0
        context['month_sales'] = Payment.objects.filter(status='completed', payment_date__gte=current_month_start).aggregate(total=Sum('amount'))['total'] or 0
        context['year_sales'] = Payment.objects.filter(status='completed', payment_date__gte=current_year_start).aggregate(total=Sum('amount'))['total'] or 0

        context['total_quotations'] = Quotation.objects.filter(is_deleted=False).count()
        context['pending_quotations'] = Quotation.objects.filter(is_deleted=False, status='sent').count()
        context['approved_quotations'] = Quotation.objects.filter(is_deleted=False, status='approved').count()
        context['rejected_quotations'] = Quotation.objects.filter(is_deleted=False, status='rejected').count()

        context['total_invoices'] = Invoice.objects.filter(is_deleted=False).count()
        context['paid_invoices'] = Invoice.objects.filter(is_deleted=False, status='paid').count()
        context['unpaid_invoices'] = Invoice.objects.filter(is_deleted=False, status='unpaid').count()
        context['overdue_invoices'] = Invoice.objects.filter(is_deleted=False, status__in=['unpaid', 'partial_paid'], due_date__lt=today).count()

        context['total_payments'] = Payment.objects.filter(is_deleted=False, status='completed').count()
        context['total_received'] = Payment.objects.filter(is_deleted=False, status='completed').aggregate(total=Sum('amount'))['total'] or 0
        invoices = Invoice.objects.filter(is_deleted=False, status__in=['unpaid', 'partial_paid']).select_related('client')
        context['outstanding_amount'] = sum(inv.balance_due for inv in invoices)

        context['total_clients'] = Client.objects.filter(is_deleted=False).count()
        context['active_clients'] = Client.objects.filter(is_deleted=False).annotate(
            total_spent=Sum('invoices__paid_amount')
        ).filter(total_spent__gt=0).count()

        context['recent_reports'] = SavedReport.objects.select_related('generated_by', 'template')[:10]
        return context


class BaseReportView(LoginRequiredMixin, RoleRequiredMixin, PermissionRequiredMixin, ListViewMixin, ListView):
    template_name = 'reports/report_base.html'
    context_object_name = 'data'
    required_roles = ['Super Admin', 'Admin']
    required_permission = 'reports.view_reports'
    paginate_by = 50
    report_title = 'Report'
    report_type = 'general'

    def get_queryset(self):
        return []

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['report_title'] = self.report_title
        context['report_type'] = self.report_type
        context['export_formats'] = ['pdf', 'excel']
        return context


class SalesReportView(BaseReportView):
    template_name = 'reports/sales_report.html'
    report_title = 'Sales Report'
    report_type = 'sales'

    def get_queryset(self):
        queryset = Payment.objects.filter(is_deleted=False, status='completed').select_related('invoice__client')
        report_period = self.request.GET.get('period', 'monthly')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        today = timezone.now().date()

        if report_period == 'daily' and not date_from:
            date_from = today
            date_to = today
        elif report_period == 'monthly' and not date_from:
            date_from = today.replace(day=1)
            date_to = today
        elif report_period == 'yearly' and not date_from:
            date_from = today.replace(month=1, day=1)
            date_to = today

        if date_from:
            queryset = queryset.filter(payment_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(payment_date__lte=date_to)

        return queryset.order_by('-payment_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        context['total_amount'] = queryset.aggregate(total=Sum('amount'))['total'] or 0
        context['report_period'] = self.request.GET.get('period', 'monthly')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['payment_methods_summary'] = queryset.values('payment_method_name').annotate(
            count=Count('id'), total=Sum('amount')
        ).order_by('-total')
        return context


class QuotationReportView(BaseReportView):
    template_name = 'reports/quotation_report.html'
    report_title = 'Quotation Report'
    report_type = 'quotation'

    def get_queryset(self):
        queryset = Quotation.objects.filter(is_deleted=False).select_related('client')
        status_filter = self.request.GET.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        status_filter = self.request.GET.get('status', '')
        base_qs = Quotation.objects.filter(is_deleted=False)
        if status_filter:
            base_qs = base_qs.filter(status=status_filter)
        context['status_filter'] = status_filter
        context['total_count'] = base_qs.count()
        context['approved_count'] = base_qs.filter(status='approved').count()
        context['rejected_count'] = base_qs.filter(status='rejected').count()
        context['pending_count'] = base_qs.filter(status='sent').count()
        return context


class InvoiceReportView(BaseReportView):
    template_name = 'reports/invoice_report.html'
    report_title = 'Invoice Report'
    report_type = 'invoice'

    def get_queryset(self):
        queryset = Invoice.objects.filter(is_deleted=False).select_related('client')
        status_filter = self.request.GET.get('status')
        if status_filter == 'overdue':
            queryset = queryset.filter(status__in=['unpaid', 'partial_paid'], due_date__lt=timezone.now().date())
        elif status_filter:
            queryset = queryset.filter(status=status_filter)
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        status_filter = self.request.GET.get('status', '')
        base_qs = Invoice.objects.filter(is_deleted=False)
        if status_filter == 'overdue':
            base_qs = base_qs.filter(status__in=['unpaid', 'partial_paid'], due_date__lt=timezone.now().date())
        elif status_filter:
            base_qs = base_qs.filter(status=status_filter)
        context['status_filter'] = status_filter
        context['total_count'] = base_qs.count()
        context['paid_count'] = base_qs.filter(status='paid').count()
        context['unpaid_count'] = base_qs.filter(status='unpaid').count()
        context['overdue_count'] = Invoice.objects.filter(is_deleted=False, status__in=['unpaid', 'partial_paid'], due_date__lt=timezone.now().date()).count()
        context['total_amount'] = sum(inv.total_amount for inv in base_qs.iterator())
        context['paid_amount'] = base_qs.aggregate(total=Sum('paid_amount'))['total'] or 0
        context['due_amount'] = context['total_amount'] - context['paid_amount']
        return context


class PaymentReportView(BaseReportView):
    template_name = 'reports/payment_report.html'
    report_title = 'Payment Report'
    report_type = 'payment'

    def get_queryset(self):
        queryset = Payment.objects.filter(is_deleted=False).select_related('invoice__client', 'payment_method')
        status_filter = self.request.GET.get('status')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if date_from:
            queryset = queryset.filter(payment_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(payment_date__lte=date_to)
        return queryset.order_by('-payment_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        context['total_received'] = queryset.filter(status='completed').aggregate(total=Sum('amount'))['total'] or 0
        context['total_pending'] = queryset.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0
        context['total_failed'] = queryset.filter(status='failed').aggregate(total=Sum('amount'))['total'] or 0
        invoices = Invoice.objects.filter(is_deleted=False, status__in=['unpaid', 'partial_paid']).select_related('client')
        context['outstanding'] = sum(inv.balance_due for inv in invoices)
        context['status_filter'] = self.request.GET.get('status', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        return context


class ClientReportView(BaseReportView):
    template_name = 'reports/client_report.html'
    report_title = 'Client Report'
    report_type = 'client'
    paginate_by = 100

    def get_queryset(self):
        return Client.objects.filter(is_deleted=False).annotate(
            total_spent=Sum('invoices__paid_amount'),
            invoice_count=Count('invoices', distinct=True),
            quotation_count=Count('quotations', distinct=True),
        ).order_by('-total_spent')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_clients'] = Client.objects.filter(is_deleted=False).count()
        context['active_clients'] = Client.objects.filter(is_deleted=False).annotate(
            total_spent=Sum('invoices__paid_amount')
        ).filter(total_spent__gt=0).count()
        top = self.get_queryset()[:10]
        context['top_clients'] = top
        context['top_clients_labels'] = [c.company_name for c in top]
        context['top_clients_data'] = [float(c.total_spent or 0) for c in top]
        return context


def _build_pdf_report(title, headers, rows, filename, summary=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm, topMargin=15*mm, bottomGap=15*mm)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=20, alignment=TA_CENTER)
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 10))

    if summary:
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%B %d, %Y %H:%M:%S')}", styles['Normal']))
        elements.append(Spacer(1, 10))

    data = [headers]
    for row in rows:
        data.append([str(cell) if cell is not None else '' for cell in row])

    table = Table(data, colWidths=[80*mm, *(60*mm for _ in headers[1:])], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2c6fbb')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))

    if summary:
        elements.append(Paragraph("Summary", styles['Heading2']))
        for k, v in summary.items():
            elements.append(Paragraph(f"<b>{k}:</b> {v}", styles['Normal']))
        elements.append(Spacer(1, 10))

    elements.append(Paragraph(f"Report generated by QI Manager on {timezone.now().strftime('%B %d, %Y %H:%M:%S')}", styles['Normal']))
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


class BaseExportView(LoginRequiredMixin, RoleRequiredMixin, PermissionRequiredMixin, View):
    required_roles = ['Super Admin', 'Admin']
    required_permission = 'reports.export_report'

    def export_as_excel(self, headers, rows, title, filename, summary=None):
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2c6fbb', end_color='2c6fbb', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(vertical='top', wrap_text=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else '')
                cell.alignment = cell_alignment

        for col_num in range(1, len(headers) + 1):
            max_length = max(len(str(headers[col_num - 1])), 10)
            for row_num in range(2, len(rows) + 2):
                cell_value = ws.cell(row=row_num, column=col_num).value or ''
                max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)

        if summary:
            start_row = len(rows) + 3
            for k, v in summary.items():
                cell = ws.cell(row=start_row, column=1, value=f"{k}:")
                cell.font = Font(bold=True)
                ws.cell(row=start_row, column=2, value=str(v))
                start_row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response


class SalesExportView(BaseExportView):
    def get(self, request):
        queryset = Payment.objects.filter(is_deleted=False, status='completed').select_related('invoice__client').order_by('-payment_date')
        report_period = request.GET.get('period', 'monthly')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        today = timezone.now().date()

        if report_period == 'daily' and not date_from:
            date_from = today
            date_to = today
        elif report_period == 'monthly' and not date_from:
            date_from = today.replace(day=1)
            date_to = today
        elif report_period == 'yearly' and not date_from:
            date_from = today.replace(month=1, day=1)
            date_to = today

        if date_from:
            queryset = queryset.filter(payment_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(payment_date__lte=date_to)

        headers = ['#', 'Payment Date', 'Invoice Number', 'Client', 'Payment Method', 'Reference', 'Amount', 'Status']
        rows = []
        for i, p in enumerate(queryset, 1):
            rows.append([i, p.payment_date.strftime('%Y-%m-%d'), p.invoice.invoice_number, p.invoice.client.company_name, p.payment_method_name, p.reference_number or '-', f"${p.amount:.2f}", p.get_status_display()])

        total = queryset.aggregate(total=Sum('amount'))['total'] or 0
        summary = {'Total Amount': f"${total:.2f}", 'Record Count': queryset.count(), 'Period': f"{date_from} to {date_to}"}
        format_type = request.GET.get('format', 'excel')
        if format_type == 'pdf':
            pdf = _build_pdf_report('Sales Report', headers, rows, 'sales_report.pdf', summary)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'
            return response
        return self.export_as_excel(headers, rows, 'Sales Report', f'sales_report_{timezone.now().strftime("%Y%m%d")}', summary)


class QuotationExportView(BaseExportView):
    def get(self, request):
        status_filter = request.GET.get('status', '')
        queryset = Quotation.objects.filter(is_deleted=False).select_related('client')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        headers = ['#', 'Quotation Number', 'Client', 'Contact Person', 'Status', 'Total Amount', 'Created Date', 'Valid Until']
        rows = []
        for i, q in enumerate(queryset.order_by('-created_at'), 1):
            rows.append([i, q.quotation_number, q.client.company_name, q.contact_person.person_name if q.contact_person else '-', q.get_status_display(), f"${q.total_amount:.2f}", q.created_at.strftime('%Y-%m-%d'), q.valid_until.strftime('%Y-%m-%d') if q.valid_until else '-'])

        total = sum(q.total_amount for q in queryset.iterator())
        summary = {'Total Quotations': queryset.count(), 'Total Value': f"${total:.2f}", 'Filter': status_filter or 'All'}
        format_type = request.GET.get('format', 'excel')
        if format_type == 'pdf':
            pdf = _build_pdf_report('Quotation Report', headers, rows, 'quotation_report.pdf', summary)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="quotation_report.pdf"'
            return response
        return self.export_as_excel(headers, rows, 'Quotation Report', f'quotation_report_{timezone.now().strftime("%Y%m%d")}', summary)


class InvoiceExportView(BaseExportView):
    def get(self, request):
        status_filter = request.GET.get('status', '')
        queryset = Invoice.objects.filter(is_deleted=False).select_related('client')
        if status_filter == 'overdue':
            queryset = queryset.filter(status__in=['unpaid', 'partial_paid'], due_date__lt=timezone.now().date())
        elif status_filter:
            queryset = queryset.filter(status=status_filter)

        headers = ['#', 'Invoice Number', 'Client', 'Status', 'Total Amount', 'Paid Amount', 'Balance Due', 'Due Date', 'Created Date']
        rows = []
        for i, inv in enumerate(queryset.order_by('-created_at'), 1):
            rows.append([i, inv.invoice_number, inv.client.company_name, inv.get_status_display(), f"${inv.total_amount:.2f}", f"${inv.paid_amount:.2f}", f"${inv.balance_due:.2f}", inv.due_date.strftime('%Y-%m-%d') if inv.due_date else '-', inv.created_at.strftime('%Y-%m-%d')])

        total_amount = sum(inv.total_amount for inv in queryset.iterator())
        paid_amount = queryset.aggregate(total=Sum('paid_amount'))['total'] or 0
        due_amount = total_amount - paid_amount
        summary = {'Total Invoices': queryset.count(), 'Total Amount': f"${total_amount:.2f}", 'Paid Amount': f"${paid_amount:.2f}", 'Due Amount': f"${due_amount:.2f}", 'Filter': status_filter or 'All'}
        format_type = request.GET.get('format', 'excel')
        if format_type == 'pdf':
            pdf = _build_pdf_report('Invoice Report', headers, rows, 'invoice_report.pdf', summary)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="invoice_report.pdf"'
            return response
        return self.export_as_excel(headers, rows, 'Invoice Report', f'invoice_report_{timezone.now().strftime("%Y%m%d")}', summary)


class PaymentExportView(BaseExportView):
    def get(self, request):
        status_filter = request.GET.get('status', '')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        queryset = Payment.objects.filter(is_deleted=False).select_related('invoice__client')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if date_from:
            queryset = queryset.filter(payment_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(payment_date__lte=date_to)

        headers = ['#', 'Payment Date', 'Invoice Number', 'Client', 'Payment Method', 'Reference', 'Amount', 'Status']
        rows = []
        for i, p in enumerate(queryset.order_by('-payment_date'), 1):
            rows.append([i, p.payment_date.strftime('%Y-%m-%d'), p.invoice.invoice_number, p.invoice.client.company_name, p.payment_method_name, p.reference_number or '-', f"${p.amount:.2f}", p.get_status_display()])

        total = queryset.aggregate(total=Sum('amount'))['total'] or 0
        summary = {'Total Payments': queryset.count(), 'Total Amount': f"${total:.2f}", 'Filter': status_filter or 'All'}
        format_type = request.GET.get('format', 'excel')
        if format_type == 'pdf':
            pdf = _build_pdf_report('Payment Report', headers, rows, 'payment_report.pdf', summary)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="payment_report.pdf"'
            return response
        return self.export_as_excel(headers, rows, 'Payment Report', f'payment_report_{timezone.now().strftime("%Y%m%d")}', summary)


class ClientExportView(BaseExportView):
    def get(self, request):
        queryset = Client.objects.filter(is_deleted=False).annotate(
            total_spent=Sum('invoices__paid_amount'),
            invoice_count=Count('invoices', distinct=True),
            quotation_count=Count('quotations', distinct=True),
        ).order_by('-total_spent')

        headers = ['#', 'Client Name', 'Email', 'Phone', 'Total Invoices', 'Total Quotations', 'Total Spent']
        rows = []
        for i, c in enumerate(queryset, 1):
            rows.append([i, c.company_name, c.email or '-', c.phone or '-', c.invoice_count, c.quotation_count, f"${c.total_spent:.2f}" if c.total_spent else '$0.00'])

        summary = {'Total Clients': queryset.count(), 'Active Clients': queryset.filter(total_spent__gt=0).count()}
        format_type = request.GET.get('format', 'excel')
        if format_type == 'pdf':
            pdf = _build_pdf_report('Client Report', headers, rows, 'client_report.pdf', summary)
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="client_report.pdf"'
            return response
        return self.export_as_excel(headers, rows, 'Client Report', f'client_report_{timezone.now().strftime("%Y%m%d")}', summary)
